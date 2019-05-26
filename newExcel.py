import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
import datetime
import os

#Creating
wb = openpyxl.Workbook()
wb.get_sheet_names()
sheet = wb.active

sheet.title = "Spam Bacon Eggs"

# Create default name sheet in last index
wb.create_sheet()

# create sheet with arguments
wb.create_sheet(index=3, title="index 3")

# Write something in the cells
sheet = wb.get_sheet_by_name('Spam Bacon Eggs')
sheet['A1'] = "some thing"
sheet['A2'] = 32

for i in range(1,10):
    row = 4
    coord = str(get_column_letter(i)) + str(row)
    print("Coords", coord)

# Saving
file_name = str(datetime.datetime.now().strftime('%Y-%m-%d %H%M%S')) + "-CreationTest.xlsx"
print(os.getcwd())
full_path = os.path.join(os.getcwd(), "XLSX", file_name)
print(full_path)
folders_only = os.path.dirname(full_path)
print(folders_only)

if (not os.path.exists(folders_only)):
    os.makedirs(folders_only)

wb.save(full_path)
