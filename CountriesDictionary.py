import openpyxl

wb = openpyxl.load_workbook('Countries.xlsx')
sheet1 = wb['Sheet1']
maxrow = sheet1.max_row

print(sheet1.cell(row=1, column=1).value, maxrow)

file = open("CsharpCountries.txt", "w")

for i in range(2,maxrow,1):
    letterCode = sheet1.cell(row=i, column=1).value
    fullName = sheet1.cell(row=i, column=2).value
    print(i, letterCode, fullName)
    file.write(f'countryList.Add("{letterCode}", "{fullName}");\n')

file.close()