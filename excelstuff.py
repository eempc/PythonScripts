import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string


print("---- Start ----")
#Load workbook wb in cell value mode (i.e. formulas like =D2+E3 will be displayed)
wb = openpyxl.load_workbook('readtest.xlsx')
# Add the data_only argument to be get the derived value from formulas
wb2 = openpyxl.load_workbook('readtest.xlsx', data_only=True)
#print(type(wb))

# Select a sheet

sheetNamesArray = wb.get_sheet_names() #deprecated
sheetNamesArray2 = wb.sheetnames # Use this to get a list of sheet names in case you do not know it (or need to iterate through them all), you can then access it like an array/tuple

print(sheetNamesArray2)

sheet1 = wb.get_sheet_by_name('Sheet1') #Use this if you know the sheet name that you want to work with
print(type(sheet1))
print(sheet1.title)

lastActiveSheet = wb.active
print(lastActiveSheet)

targetSheetString = sheetNamesArray2[0]
print(targetSheetString) # No need for .title, it returns the name only, not a sheet object, this is string

current_worksheet = wb.get_sheet_by_name(sheetNamesArray2[0]) # is sheet object

# Get a cell value

print(current_worksheet['B2'].value)
print(current_worksheet['C2'].value)
print(current_worksheet['B2'].column)
print(current_worksheet['B2'].row)
print(current_worksheet['B2'].coordinate)

print(current_worksheet.cell(row=1, column=2).value)

# Iterating through a row
for i in range(1,4,1):
    print(i, current_worksheet.cell(row=1, column=i).value)

# Print all of column B'current_worksheet values (auto stops after reaching bounds), it'current_worksheet a tuple

print(current_worksheet.columns[1])
for cell in current_worksheet.columns[1]:
    print(cell.value)

# Get max used row and column (i.e. bounds)
print(current_worksheet.max_row)
print(current_worksheet.max_column)

# Converting between the Excel letters and integer

print(get_column_letter(27))  # Expect AA
print(column_index_from_string('C'))  # Expect 3

# Iterate through all used cells row by row, swap max_row and max_column to go through vertically
print("---iterate through all row by row---")
for i in range(1, current_worksheet.max_row, 1):
    print("--Row", i,"--")
    for j in range(1, current_worksheet.max_column, 1):
        cell = current_worksheet.cell(row=i, column=j)
        print(cell.coordinate, cell.value)
        if (cell.value == "bold"):
            print("Found", cell.column, cell.row)
            print("Down one", cell.column, cell.row + 1)
            print("Right one",get_column_letter(column_index_from_string(cell.column) + 1), cell.row)
print("---end---")

# Slice a sheet section and iterate through each cell

tupleSlice = tuple(current_worksheet['A1':'C3'])

i = 0
for row in tupleSlice:
    print("*Row " + str(i))
    for cellObj in row:
        print(cellObj.coordinate, cellObj.value)
    print ("--End Row--")
    i += 1

print("========")

# Create a function to move to return a different cell. Enter a cell object and the movement. Return a coordinate e.g. 'B3'
def shift_coordinate(cell, column_delta, row_delta):
    #print("Current cell", cell.coordinate)
    #print("Column letter", cell.column)
    #print("Column index", column_index_from_string(cell.column))
    #print("Row",cell.row)
    #print("New coordinates follows")

    new_row = cell.row + row_delta
    new_column_index = column_index_from_string(cell.column) + column_delta
    if (new_row < 1 or new_column_index < 1):
        return "Out of bounds"
    else:
        new_column_letter = get_column_letter(new_column_index)
        new_coordinates = new_column_letter + str(new_row)
        #print(new_coordinates)
        #print("Double check", current_worksheet[new_coordinates].coordinate)
        return new_coordinates

print(shift_coordinate(current_worksheet['C4'], -1, 1))

